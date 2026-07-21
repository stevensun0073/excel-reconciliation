import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import re

def highlight_two_sheets(file_path, output_path):
    print("=" * 60)
    print("【跨表双向绿色高亮对账系统】")
    print("=" * 60)
    
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"[-] 读取文件失败: {e}")
        return

    # 检查工作表数量
    sheet_names = wb.sheetnames
    if len(sheet_names) < 2:
        print(f"[-] 错误：当前文件只有 1 个工作表，找不到两个表来比对。当前表名有: {sheet_names}")
        return

    # 获取前两个工作表
    ws1 = wb[sheet_names[0]]
    ws2 = wb[sheet_names[1]]
    print(f"[+] 正在比对的两个表分别是: 【{sheet_names[0]}】 和 【{sheet_names[1]}】")

    def clean_number(val):
        if val is None:
            return None
        try:
            val_str = re.sub(r'[^\d.-]', '', str(val))
            if val_str == '' or val_str == '-' or val_str == '.':
                return None
            return float(val_str)
        except:
            return None

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    
    # 提取表1的所有有效单元格数字位置
    list1 = []
    for r in range(1, ws1.max_row + 1):
        for c in range(1, ws1.max_column + 1):
            val = clean_number(ws1.cell(row=r, column=c).value)
            if val is not None:
                list1.append((r, c, val))

    # 提取表2的所有有效单元格数字位置
    list2 = []
    for r in range(1, ws2.max_row + 1):
        for c in range(1, ws2.max_column + 1):
            val = clean_number(ws2.cell(row=r, column=c).value)
            if val is not None:
                list2.append((r, c, val))

    matched_count = 0
    used2 = set()

    # 跨表比对
    for r1, c1, v1 in list1:
        for idx, (r2, c2, v2) in enumerate(list2):
            if idx in used2:
                continue
            
            if abs(v1 - v2) <= 0.01:
                # 同时将表1和表2中对应的单元格标为绿色
                ws1.cell(row=r1, column=c1).fill = green_fill
                ws2.cell(row=r2, column=c2).fill = green_fill
                
                used2.add(idx)
                matched_count += 1
                break

    try:
        wb.save(output_path)
        print(f"[√] 跨表对账完成！共成功配对 {matched_count} 处相同数值，双方表格的对应单元格已同时标绿。")
        print(f"[√] 结果文件已保存: 【{output_path}】")
    except Exception as e:
        print(f"[-] 保存文件失败: {e}")

highlight_two_sheets('202605.xlsx', '跨表对账高亮结果.xlsx')