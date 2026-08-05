from __future__ import annotations

from collections import defaultdict
from copy import copy
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill


# ============================================================
# 文件配置
# ============================================================

INPUT_FILE = Path("result_reconciliation.xlsx")
OUTPUT_FILE = Path("result_yellow_same_reference.xlsx")

SHEET1_NAME = "Sheet1"
SHEET2_NAME = "Sheet2"

HEADER_ROW = 2
DATA_START_ROW = 3

AMOUNT_TOLERANCE = Decimal("0.01")

# 黄色
YELLOW_FILL = PatternFill(
    start_color="FFF2CC",
    end_color="FFF2CC",
    fill_type="solid",
)

MATCH_TYPE_TEXT = "Yellow-Same-Bank-Reference-to-One"


# ============================================================
# 可识别的表头名称
# ============================================================

BANK_REFERENCE_HEADERS = {
    "银行流水号",
    "银行交易流水号",
    "银行参考号",
    "流水号",
    "bank reference",
    "bank reference number",
    "reference",
    "reference number",
}

AMOUNT_HEADERS = {
    "以公司代码货币计算的金额",
    "金额",
    "交易金额",
    "本币金额",
    "amount",
    "transaction amount",
}

MATCH_TYPE_HEADERS = {
    "匹配类型",
    "match type",
}

COUNTERPART_ROW_HEADERS = {
    "对方行号",
    "匹配行号",
    "counterpart row",
    "matched row",
    "partner rows",
}

REVIEW_HEADERS = {
    "复核",
    "review",
}


# ============================================================
# 基础工具
# ============================================================

def normalize_text(value: Any) -> str:
    if value is None:
        return ""

    return str(value).strip().lower()


def normalize_reference(value: Any) -> str:
    """
    标准化银行流水号。

    当前只处理：
    - 去除首尾空格
    - 不区分大小写

    不删除中间空格、横杠或其他符号，
    避免把不同流水号错误合并。
    """
    return normalize_text(value)


def to_decimal(value: Any) -> Decimal | None:
    if value is None:
        return None

    if isinstance(value, bool):
        return None

    if isinstance(value, Decimal):
        return value

    if isinstance(value, (int, float)):
        return Decimal(str(value))

    text = str(value).strip()

    if not text:
        return None

    text = text.replace(",", "")

    if text.startswith("(") and text.endswith(")"):
        text = f"-{text[1:-1]}"

    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def amounts_equal(amount1: Decimal, amount2: Decimal) -> bool:
    return abs(amount1 - amount2) <= AMOUNT_TOLERANCE


def copy_cell_style(source_cell, target_cell) -> None:
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.fill = copy(source_cell.fill)
        target_cell.border = copy(source_cell.border)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.number_format = source_cell.number_format
        target_cell.protection = copy(source_cell.protection)


def copy_worksheet(source_ws, target_ws) -> None:
    """
    将原工作表内容和主要格式复制到新工作簿。
    """
    for row in source_ws.iter_rows():
        for source_cell in row:
            target_cell = target_ws[source_cell.coordinate]
            target_cell.value = source_cell.value

            if source_cell.has_style:
                copy_cell_style(source_cell, target_cell)

    for column_letter, dimension in source_ws.column_dimensions.items():
        target_ws.column_dimensions[column_letter].width = dimension.width
        target_ws.column_dimensions[column_letter].hidden = dimension.hidden

    for row_number, dimension in source_ws.row_dimensions.items():
        target_ws.row_dimensions[row_number].height = dimension.height
        target_ws.row_dimensions[row_number].hidden = dimension.hidden

    for merged_range in source_ws.merged_cells.ranges:
        target_ws.merge_cells(str(merged_range))

    target_ws.freeze_panes = source_ws.freeze_panes
    target_ws.sheet_view.showGridLines = source_ws.sheet_view.showGridLines

    if source_ws.auto_filter.ref:
        target_ws.auto_filter.ref = source_ws.auto_filter.ref


def find_column(
    worksheet,
    accepted_headers: set[str],
    required: bool = True,
) -> int | None:
    for cell in worksheet[HEADER_ROW]:
        header = normalize_text(cell.value)

        if header in accepted_headers:
            return cell.column

    if required:
        available_headers = [
            str(cell.value).strip()
            for cell in worksheet[HEADER_ROW]
            if cell.value is not None
        ]

        raise ValueError(
            f"在工作表 {worksheet.title!r} 中找不到表头："
            f"{sorted(accepted_headers)}。\n"
            f"当前第 {HEADER_ROW} 行表头为：{available_headers}"
        )

    return None


def find_or_create_column(
    worksheet,
    accepted_headers: set[str],
    new_header: str,
) -> int:
    existing_column = find_column(
        worksheet=worksheet,
        accepted_headers=accepted_headers,
        required=False,
    )

    if existing_column is not None:
        return existing_column

    new_column = worksheet.max_column + 1
    worksheet.cell(HEADER_ROW, new_column).value = new_header

    return new_column


def row_is_unmatched(
    worksheet,
    row_number: int,
    match_type_column: int,
    counterpart_row_column: int,
) -> bool:
    """
    只有匹配类型和对方行号都为空，才视为未匹配。

    因此原有绿色和棕色记录不会被重新参与黄色匹配。
    """
    match_type = worksheet.cell(row_number, match_type_column).value
    counterpart = worksheet.cell(row_number, counterpart_row_column).value

    return (
        normalize_text(match_type) == ""
        and normalize_text(counterpart) == ""
    )


def apply_yellow_fill(worksheet, row_number: int) -> None:
    """
    给整条数据行加黄色。

    只改变新匹配记录的填充色，不改字体、边框和数字格式。
    """
    for column_number in range(1, worksheet.max_column + 1):
        worksheet.cell(row_number, column_number).fill = copy(YELLOW_FILL)


# ============================================================
# 工作簿处理
# ============================================================

def create_output_workbook(input_workbook) -> Workbook:
    output_workbook = Workbook()

    default_sheet = output_workbook.active
    output_workbook.remove(default_sheet)

    for source_ws in input_workbook.worksheets:
        target_ws = output_workbook.create_sheet(source_ws.title)
        copy_worksheet(source_ws, target_ws)

    return output_workbook


def collect_sheet1_reference_groups(
    worksheet,
    bank_reference_column: int,
    amount_column: int,
    match_type_column: int,
    counterpart_row_column: int,
) -> dict[str, list[tuple[int, Decimal]]]:
    """
    收集 Sheet1 中仍未匹配的记录，并按银行流水号分组。
    """
    groups: dict[str, list[tuple[int, Decimal]]] = defaultdict(list)

    for row_number in range(DATA_START_ROW, worksheet.max_row + 1):
        if not row_is_unmatched(
            worksheet,
            row_number,
            match_type_column,
            counterpart_row_column,
        ):
            continue

        raw_reference = worksheet.cell(
            row_number,
            bank_reference_column,
        ).value

        reference = normalize_reference(raw_reference)

        if not reference:
            continue

        amount = to_decimal(
            worksheet.cell(row_number, amount_column).value
        )

        if amount is None:
            continue

        groups[reference].append((row_number, amount))

    return groups


def collect_unmatched_sheet2_rows(
    worksheet,
    amount_column: int,
    match_type_column: int,
    counterpart_row_column: int,
) -> list[tuple[int, Decimal]]:
    unmatched_rows: list[tuple[int, Decimal]] = []

    for row_number in range(DATA_START_ROW, worksheet.max_row + 1):
        if not row_is_unmatched(
            worksheet,
            row_number,
            match_type_column,
            counterpart_row_column,
        ):
            continue

        amount = to_decimal(
            worksheet.cell(row_number, amount_column).value
        )

        if amount is None:
            continue

        unmatched_rows.append((row_number, amount))

    return unmatched_rows


def perform_same_reference_group_matching(
    sheet1,
    sheet2,
    sheet1_reference_column: int,
    sheet1_amount_column: int,
    sheet2_amount_column: int,
    sheet1_match_type_column: int,
    sheet2_match_type_column: int,
    sheet1_counterpart_column: int,
    sheet2_counterpart_column: int,
    sheet1_review_column: int,
    sheet2_review_column: int,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """
    黄色阶段第一条规则：

    Sheet1 中同一银行流水号的所有未匹配行必须整体参加匹配。
    它们的金额合计，匹配 Sheet2 中一条未匹配记录。

    安全限制：
    - 至少两条 Sheet1 记录才进入本规则；
    - 不允许只选择同流水号中的部分行；
    - 如果 Sheet2 有多条相同金额候选，则不自动匹配；
    - 已匹配的绿色、棕色、旧黄色记录完全不参与。
    """
    reference_groups = collect_sheet1_reference_groups(
        worksheet=sheet1,
        bank_reference_column=sheet1_reference_column,
        amount_column=sheet1_amount_column,
        match_type_column=sheet1_match_type_column,
        counterpart_row_column=sheet1_counterpart_column,
    )

    sheet2_unmatched = collect_unmatched_sheet2_rows(
        worksheet=sheet2,
        amount_column=sheet2_amount_column,
        match_type_column=sheet2_match_type_column,
        counterpart_row_column=sheet2_counterpart_column,
    )

    used_sheet2_rows: set[int] = set()

    successful_matches: list[dict[str, Any]] = []
    ambiguous_candidates: list[dict[str, Any]] = []

    # 优先处理行数较多的组，再按最早行号排序。
    ordered_groups = sorted(
        reference_groups.items(),
        key=lambda item: (
            -len(item[1]),
            min(row_number for row_number, _ in item[1]),
        ),
    )

    for normalized_reference, group_rows in ordered_groups:
        # 这条规则只处理“多条对一条”。
        if len(group_rows) < 2:
            continue

        sheet1_row_numbers = [
            row_number
            for row_number, _ in group_rows
        ]

        sheet1_total = sum(
            (amount for _, amount in group_rows),
            Decimal("0.00"),
        )

        candidate_sheet2_rows = [
            row_number
            for row_number, amount in sheet2_unmatched
            if row_number not in used_sheet2_rows
            and amounts_equal(sheet1_total, amount)
        ]

        # 没有金额一致的单条 Sheet2 记录。
        if not candidate_sheet2_rows:
            continue

        # 出现多个同金额候选时不自动选择，避免错误匹配。
        if len(candidate_sheet2_rows) > 1:
            ambiguous_candidates.append(
                {
                    "reference": normalized_reference,
                    "sheet1_rows": sheet1_row_numbers,
                    "sheet1_count": len(sheet1_row_numbers),
                    "sheet1_total": sheet1_total,
                    "sheet2_candidates": candidate_sheet2_rows,
                }
            )
            continue

        sheet2_row_number = candidate_sheet2_rows[0]
        used_sheet2_rows.add(sheet2_row_number)

        sheet1_rows_text = ", ".join(
            str(row_number)
            for row_number in sheet1_row_numbers
        )

        # 更新 Sheet1：同流水号的所有未匹配行全部匹配同一条 Sheet2。
        for sheet1_row_number in sheet1_row_numbers:
            sheet1.cell(
                sheet1_row_number,
                sheet1_match_type_column,
            ).value = MATCH_TYPE_TEXT

            sheet1.cell(
                sheet1_row_number,
                sheet1_counterpart_column,
            ).value = sheet2_row_number

            sheet1.cell(
                sheet1_row_number,
                sheet1_review_column,
            ).value = "同银行流水号全部记录金额合计匹配Sheet2单条"

            apply_yellow_fill(sheet1, sheet1_row_number)

        # 更新 Sheet2：记录对应的全部 Sheet1 行号。
        sheet2.cell(
            sheet2_row_number,
            sheet2_match_type_column,
        ).value = MATCH_TYPE_TEXT

        sheet2.cell(
            sheet2_row_number,
            sheet2_counterpart_column,
        ).value = sheet1_rows_text

        sheet2.cell(
            sheet2_row_number,
            sheet2_review_column,
        ).value = "Sheet1同银行流水号全部记录金额合计"

        apply_yellow_fill(sheet2, sheet2_row_number)

        successful_matches.append(
            {
                "reference": normalized_reference,
                "sheet1_rows": sheet1_row_numbers,
                "sheet1_count": len(sheet1_row_numbers),
                "sheet1_total": sheet1_total,
                "sheet2_row": sheet2_row_number,
            }
        )

    return successful_matches, ambiguous_candidates


def write_summary_sheet(
    workbook,
    successful_matches: list[dict[str, Any]],
    ambiguous_candidates: list[dict[str, Any]],
) -> None:
    summary_name = "Yellow Match Summary"

    if summary_name in workbook.sheetnames:
        del workbook[summary_name]

    worksheet = workbook.create_sheet(summary_name)

    worksheet.append(
        [
            "Status",
            "Bank Reference",
            "Sheet1 Rows",
            "Sheet1 Count",
            "Sheet1 Total",
            "Sheet2 Row / Candidates",
            "Explanation",
        ]
    )

    for match in successful_matches:
        worksheet.append(
            [
                "Matched",
                match["reference"],
                ", ".join(str(row) for row in match["sheet1_rows"]),
                match["sheet1_count"],
                float(match["sheet1_total"]),
                match["sheet2_row"],
                "同银行流水号的全部未匹配记录，合计匹配Sheet2单条记录",
            ]
        )

    for candidate in ambiguous_candidates:
        worksheet.append(
            [
                "Ambiguous - Not Matched",
                candidate["reference"],
                ", ".join(str(row) for row in candidate["sheet1_rows"]),
                candidate["sheet1_count"],
                float(candidate["sheet1_total"]),
                ", ".join(
                    str(row)
                    for row in candidate["sheet2_candidates"]
                ),
                "Sheet2存在多条相同金额候选，未自动选择",
            ]
        )

    worksheet.freeze_panes = "A2"

    column_widths = {
        "A": 24,
        "B": 28,
        "C": 34,
        "D": 14,
        "E": 18,
        "F": 30,
        "G": 60,
    }

    for column_letter, width in column_widths.items():
        worksheet.column_dimensions[column_letter].width = width


def main() -> None:
    if not INPUT_FILE.exists():
        raise FileNotFoundError(
            f"找不到输入文件：{INPUT_FILE.resolve()}"
        )

    print("=" * 72)
    print("Yellow Same Bank Reference Matching")
    print("=" * 72)
    print(f"Input : {INPUT_FILE.resolve()}")
    print(f"Output: {OUTPUT_FILE.resolve()}")
    print()

    input_workbook = openpyxl.load_workbook(INPUT_FILE)
    output_workbook = create_output_workbook(input_workbook)

    if SHEET1_NAME not in output_workbook.sheetnames:
        raise ValueError(f"找不到工作表：{SHEET1_NAME}")

    if SHEET2_NAME not in output_workbook.sheetnames:
        raise ValueError(f"找不到工作表：{SHEET2_NAME}")

    sheet1 = output_workbook[SHEET1_NAME]
    sheet2 = output_workbook[SHEET2_NAME]

    # Sheet1 列
    sheet1_reference_column = find_column(
        sheet1,
        BANK_REFERENCE_HEADERS,
    )

    sheet1_amount_column = find_column(
        sheet1,
        AMOUNT_HEADERS,
    )

    sheet1_match_type_column = find_or_create_column(
        sheet1,
        MATCH_TYPE_HEADERS,
        "匹配类型",
    )

    sheet1_counterpart_column = find_or_create_column(
        sheet1,
        COUNTERPART_ROW_HEADERS,
        "对方行号",
    )

    sheet1_review_column = find_or_create_column(
        sheet1,
        REVIEW_HEADERS,
        "复核",
    )

    # Sheet2 列
    sheet2_amount_column = find_column(
        sheet2,
        AMOUNT_HEADERS,
    )

    sheet2_match_type_column = find_or_create_column(
        sheet2,
        MATCH_TYPE_HEADERS,
        "匹配类型",
    )

    sheet2_counterpart_column = find_or_create_column(
        sheet2,
        COUNTERPART_ROW_HEADERS,
        "对方行号",
    )

    sheet2_review_column = find_or_create_column(
        sheet2,
        REVIEW_HEADERS,
        "复核",
    )

    print("Detected columns:")
    print(
        f"  Sheet1 bank reference: "
        f"{sheet1.cell(HEADER_ROW, sheet1_reference_column).value}"
    )
    print(
        f"  Sheet1 amount        : "
        f"{sheet1.cell(HEADER_ROW, sheet1_amount_column).value}"
    )
    print(
        f"  Sheet2 amount        : "
        f"{sheet2.cell(HEADER_ROW, sheet2_amount_column).value}"
    )
    print()

    successful_matches, ambiguous_candidates = (
        perform_same_reference_group_matching(
            sheet1=sheet1,
            sheet2=sheet2,
            sheet1_reference_column=sheet1_reference_column,
            sheet1_amount_column=sheet1_amount_column,
            sheet2_amount_column=sheet2_amount_column,
            sheet1_match_type_column=sheet1_match_type_column,
            sheet2_match_type_column=sheet2_match_type_column,
            sheet1_counterpart_column=sheet1_counterpart_column,
            sheet2_counterpart_column=sheet2_counterpart_column,
            sheet1_review_column=sheet1_review_column,
            sheet2_review_column=sheet2_review_column,
        )
    )

    write_summary_sheet(
        workbook=output_workbook,
        successful_matches=successful_matches,
        ambiguous_candidates=ambiguous_candidates,
    )

    output_workbook.save(OUTPUT_FILE)

    matched_sheet1_rows = sum(
        match["sheet1_count"]
        for match in successful_matches
    )

    print("=" * 72)
    print("Result")
    print("=" * 72)
    print(f"Successful reference groups : {len(successful_matches)}")
    print(f"Matched Sheet1 rows         : {matched_sheet1_rows}")
    print(f"Matched Sheet2 rows         : {len(successful_matches)}")
    print(f"Ambiguous groups not matched: {len(ambiguous_candidates)}")
    print()
    print(f"Saved: {OUTPUT_FILE.resolve()}")

    if successful_matches:
        print()
        print("Matched groups:")

        for match in successful_matches:
            print(
                f"  Reference={match['reference']!r}, "
                f"Sheet1 rows={match['sheet1_rows']}, "
                f"total={match['sheet1_total']}, "
                f"Sheet2 row={match['sheet2_row']}"
            )

    if ambiguous_candidates:
        print()
        print("Ambiguous groups left unmatched:")

        for candidate in ambiguous_candidates:
            print(
                f"  Reference={candidate['reference']!r}, "
                f"Sheet1 rows={candidate['sheet1_rows']}, "
                f"total={candidate['sheet1_total']}, "
                f"Sheet2 candidates={candidate['sheet2_candidates']}"
            )


if __name__ == "__main__":
    main()