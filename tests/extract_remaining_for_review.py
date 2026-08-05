from __future__ import annotations

import re
from copy import copy
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill


INPUT_FILE = Path("result_reconciliation.xlsx")
OUTPUT_FILE = Path("remaining_for_review.xlsx")

SHEET_NAMES = ["Sheet1", "Sheet2"]
HEADER_ROW = 2
DATA_START_ROW = 3

NO_FILL = PatternFill(fill_type=None)


MATCH_TYPE_HEADERS = {
    "match type",
    "匹配类型",
}


def normalize_text(value: Any) -> str:
    if value is None:
        return ""

    return str(value).strip().lower()


def find_match_type_column(worksheet) -> int:
    for cell in worksheet[HEADER_ROW]:
        if normalize_text(cell.value) in MATCH_TYPE_HEADERS:
            return cell.column

    headers = [
        str(cell.value).strip()
        for cell in worksheet[HEADER_ROW]
        if cell.value is not None
    ]

    raise ValueError(
        f"工作表 {worksheet.title!r} 找不到 Match Type / 匹配类型列。\n"
        f"当前表头：{headers}"
    )


def is_confirmed_green_or_brown(match_type: Any) -> bool:
    """
    判断是否属于已经确认的绿色或棕色匹配。

    可识别：
    One-to-One
    One-to-Two
    Two-to-One
    Three-to-One
    Four-to-One
    Five-to-One
    Six-to-One
    Seven-to-One
    以及相同结构的其他数字组合。

    不删除：
    Keyword Difference
    Keyword-Difference (...)
    Final Yellow Match (...)
    空白未匹配记录
    """
    text = normalize_text(match_type)

    if not text:
        return False

    number_words = (
        r"one|two|three|four|five|six|seven|eight|nine|ten"
    )

    standard_match_pattern = re.compile(
        rf"^({number_words})-to-({number_words})$",
        re.IGNORECASE,
    )

    return standard_match_pattern.fullmatch(text) is not None


def copy_cell_style(source_cell, target_cell) -> None:
    if not source_cell.has_style:
        return

    target_cell.font = copy(source_cell.font)
    target_cell.fill = copy(source_cell.fill)
    target_cell.border = copy(source_cell.border)
    target_cell.alignment = copy(source_cell.alignment)
    target_cell.number_format = source_cell.number_format
    target_cell.protection = copy(source_cell.protection)


def copy_row(
    source_ws,
    source_row: int,
    target_ws,
    target_row: int,
    add_source_row: bool = True,
) -> None:
    """
    复制一整行的值和格式。

    输出文件第一列新增 Original Row，
    原工作表内容从第二列开始。
    """
    column_offset = 1 if add_source_row else 0

    if add_source_row:
        target_ws.cell(target_row, 1).value = source_row

    for source_column in range(1, source_ws.max_column + 1):
        source_cell = source_ws.cell(source_row, source_column)
        target_cell = target_ws.cell(
            target_row,
            source_column + column_offset,
        )

        target_cell.value = source_cell.value
        copy_cell_style(source_cell, target_cell)


def copy_column_widths(source_ws, target_ws) -> None:
    target_ws.column_dimensions["A"].width = 14

    for source_column in range(1, source_ws.max_column + 1):
        source_letter = openpyxl.utils.get_column_letter(source_column)
        target_letter = openpyxl.utils.get_column_letter(source_column + 1)

        width = source_ws.column_dimensions[source_letter].width

        if width is not None:
            target_ws.column_dimensions[target_letter].width = width


def create_review_sheet(source_ws, target_ws) -> dict[str, int]:
    match_type_column = find_match_type_column(source_ws)

    # 保留原工作表第1行。
    copy_row(
        source_ws=source_ws,
        source_row=1,
        target_ws=target_ws,
        target_row=1,
        add_source_row=True,
    )

    # 保留第2行表头，并加入原始行号表头。
    copy_row(
        source_ws=source_ws,
        source_row=HEADER_ROW,
        target_ws=target_ws,
        target_row=HEADER_ROW,
        add_source_row=True,
    )

    target_ws.cell(HEADER_ROW, 1).value = "Original Row"

    kept_count = 0
    removed_count = 0
    unmatched_count = 0
    keyword_count = 0
    old_yellow_count = 0
    other_count = 0

    target_row = DATA_START_ROW

    for source_row in range(DATA_START_ROW, source_ws.max_row + 1):
        match_type = source_ws.cell(
            source_row,
            match_type_column,
        ).value

        normalized_match_type = normalize_text(match_type)

        if is_confirmed_green_or_brown(match_type):
            removed_count += 1
            continue

        copy_row(
            source_ws=source_ws,
            source_row=source_row,
            target_ws=target_ws,
            target_row=target_row,
            add_source_row=True,
        )

        kept_count += 1

        if not normalized_match_type:
            unmatched_count += 1

        elif normalized_match_type.startswith("final yellow match"):
            old_yellow_count += 1

        elif (
            normalized_match_type.startswith("keyword difference")
            or normalized_match_type.startswith("keyword-difference")
        ):
            keyword_count += 1

        else:
            other_count += 1

        target_row += 1

    copy_column_widths(source_ws, target_ws)

    target_ws.freeze_panes = "A3"
    target_ws.auto_filter.ref = (
        f"A{HEADER_ROW}:"
        f"{openpyxl.utils.get_column_letter(target_ws.max_column)}"
        f"{target_ws.max_row}"
    )

    return {
        "kept": kept_count,
        "removed": removed_count,
        "unmatched": unmatched_count,
        "keyword": keyword_count,
        "old_yellow": old_yellow_count,
        "other": other_count,
    }


def write_summary_sheet(
    workbook: Workbook,
    statistics: dict[str, dict[str, int]],
) -> None:
    worksheet = workbook.create_sheet("Review Summary", 0)

    worksheet.append(
        [
            "Sheet",
            "Green/Brown Removed",
            "Rows Kept",
            "Unmatched",
            "Keyword Difference",
            "Old Final Yellow",
            "Other",
        ]
    )

    for sheet_name, values in statistics.items():
        worksheet.append(
            [
                sheet_name,
                values["removed"],
                values["kept"],
                values["unmatched"],
                values["keyword"],
                values["old_yellow"],
                values["other"],
            ]
        )

    worksheet.column_dimensions["A"].width = 18
    worksheet.column_dimensions["B"].width = 22
    worksheet.column_dimensions["C"].width = 14
    worksheet.column_dimensions["D"].width = 14
    worksheet.column_dimensions["E"].width = 22
    worksheet.column_dimensions["F"].width = 20
    worksheet.column_dimensions["G"].width = 14

    worksheet.freeze_panes = "A2"


def main() -> None:
    if not INPUT_FILE.exists():
        raise FileNotFoundError(
            f"找不到输入文件：{INPUT_FILE.resolve()}"
        )

    print("=" * 72)
    print("Extract Remaining Records for Review")
    print("=" * 72)
    print(f"Input : {INPUT_FILE.resolve()}")
    print(f"Output: {OUTPUT_FILE.resolve()}")
    print()

    input_workbook = openpyxl.load_workbook(INPUT_FILE)

    output_workbook = Workbook()
    output_workbook.remove(output_workbook.active)

    statistics: dict[str, dict[str, int]] = {}

    for sheet_name in SHEET_NAMES:
        if sheet_name not in input_workbook.sheetnames:
            raise ValueError(f"找不到工作表：{sheet_name}")

        source_ws = input_workbook[sheet_name]
        target_ws = output_workbook.create_sheet(sheet_name)

        sheet_statistics = create_review_sheet(
            source_ws=source_ws,
            target_ws=target_ws,
        )

        statistics[sheet_name] = sheet_statistics

    write_summary_sheet(
        workbook=output_workbook,
        statistics=statistics,
    )

    output_workbook.save(OUTPUT_FILE)

    print("Result:")
    print()

    for sheet_name, values in statistics.items():
        print(sheet_name)
        print(f"  Green/Brown removed : {values['removed']}")
        print(f"  Rows kept           : {values['kept']}")
        print(f"  Unmatched           : {values['unmatched']}")
        print(f"  Keyword Difference  : {values['keyword']}")
        print(f"  Old Final Yellow    : {values['old_yellow']}")
        print(f"  Other               : {values['other']}")
        print()

    print(f"Saved: {OUTPUT_FILE.resolve()}")


if __name__ == "__main__":
    main()