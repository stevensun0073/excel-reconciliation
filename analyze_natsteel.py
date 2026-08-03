# 修正版 analyze_natsteel.py（列号固定版）

from openpyxl import load_workbook

print("="*60)
print("NATSTEEL Analysis")
print("="*60)

wb = load_workbook("data.xlsx", data_only=True)
ws1 = wb["Sheet1"]
ws2 = wb["Sheet2"]

START_ROW = 3

# 固定列号（不再依赖表头名称）
S1_TEXT = 2
S1_AMOUNT = 3
S1_KEYWORD = 4

S2_NAME = 3      # Recipient's Account Name 所在列，如有不同请改这里
S2_AMOUNT = 6
S2_KEYWORD = 7

TARGET = "NATSTEEL"

def kw_set(v):
    if v is None:
        return set()
    return {x.strip().upper() for x in str(v).split(";") if x.strip()}

sheet1 = []
sheet2 = []

for r in range(START_ROW, ws1.max_row + 1):
    amt = ws1.cell(r, S1_AMOUNT).value
    if amt is None:
        continue
    kws = kw_set(ws1.cell(r, S1_KEYWORD).value)
    sheet1.append({
        "row": r,
        "amount": float(amt),
        "keyword": kws,
        "keyword_raw": ws1.cell(r, S1_KEYWORD).value,
        "text": ws1.cell(r, S1_TEXT).value,
    })

for r in range(START_ROW, ws2.max_row + 1):
    amt = ws2.cell(r, S2_AMOUNT).value
    if amt is None:
        continue
    kws = kw_set(ws2.cell(r, S2_KEYWORD).value)
    sheet2.append({
        "row": r,
        "amount": float(amt),
        "keyword": kws,
        "keyword_raw": ws2.cell(r, S2_KEYWORD).value,
        "text": ws2.cell(r, S2_NAME).value,
    })

s1_nat = [x for x in sheet1 if TARGET in x["keyword"]]
s2_nat = [x for x in sheet2 if TARGET in x["keyword"]]

sum1 = sum(x["amount"] for x in s1_nat)
sum2 = sum(x["amount"] for x in s2_nat)

print(f"Sheet1 NATSTEEL rows : {len(s1_nat)}")
print(f"Sheet2 NATSTEEL rows : {len(s2_nat)}")
print(f"Sheet1 total         : {sum1:.2f}")
print(f"Sheet2 total         : {sum2:.2f}")
print(f"Difference           : {sum2-sum1:.2f}")

print("\nFirst 20 Sheet1 NATSTEEL")
for x in s1_nat[:20]:
    print(x["row"], x["amount"], x["keyword_raw"], x["text"])

print("\nFirst 20 Sheet2 NATSTEEL")
for x in s2_nat[:20]:
    print(x["row"], x["amount"], x["keyword_raw"], x["text"])

print("="*60)
print("下一步再做组合搜索。")