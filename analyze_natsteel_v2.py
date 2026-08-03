"""
analyze_natsteel_v2.py

NATSTEEL 未匹配专项分析。

逻辑
----
1. 读取 result_reconciliation.xlsx。
2. 只看尚未匹配（Match Type 为空）的记录。
3. Sheet2：
   汇总尚未匹配且 Key word 为 NATSTEEL 的记录。
4. Sheet1：
   汇总尚未匹配且 Key word 包含 NATSTEEL 的记录。
5. 计算需要补充的差额：

       差额 =
       Sheet2 未匹配 NATSTEEL 合计
       - Sheet1 未匹配 NATSTEEL 合计

6. 只在 Sheet1 剩余黄色记录中搜索：
   - Match Type 为空；
   - Key word 不包含 NATSTEEL；
   - 组合金额严格等于差额。
7. 最多搜索 10 条记录。
8. 不修改任何 Excel 文件，只在终端输出候选。
"""

from __future__ import annotations

from decimal import Decimal, InvalidOperation
from pathlib import Path
from time import monotonic

from openpyxl import load_workbook


RESULT_FILE = Path("result_reconciliation.xlsx")

START_ROW = 3

# result_reconciliation.xlsx 当前列位置
S1_TEXT = 2
S1_AMOUNT = 3
S1_KEYWORD = 4
S1_MATCH_TYPE = 5

S2_NAME = 3
S2_AMOUNT = 6
S2_KEYWORD = 7
S2_MATCH_TYPE = 8

TARGET_KEYWORD = "NATSTEEL"

MAX_COMBINATION_SIZE = 10
SEARCH_TIME_LIMIT_SECONDS = 120
MAX_STATES_PER_SIZE = 1_000_000


def normalize_text(value) -> str:
    if value is None:
        return ""

    return " ".join(str(value).strip().split())


def keyword_set(value) -> set[str]:
    if value is None:
        return set()

    return {
        part.strip().upper()
        for part in str(value).split(";")
        if part.strip()
    }


def parse_amount(value) -> Decimal | None:
    if value is None:
        return None

    try:
        cleaned = str(value).replace(",", "").strip()

        if not cleaned:
            return None

        return Decimal(cleaned)

    except (InvalidOperation, ValueError, TypeError):
        return None


def is_blank(value) -> bool:
    return value is None or not str(value).strip()


def amount_to_cents(amount: Decimal) -> int:
    """把金额转成整数分，避免浮点误差。"""
    return int(amount * 100)


def cents_to_amount(cents: int) -> Decimal:
    return Decimal(cents) / Decimal(100)


def load_sheet1(worksheet) -> list[dict]:
    records = []

    for row in range(START_ROW, worksheet.max_row + 1):
        amount = parse_amount(
            worksheet.cell(row, S1_AMOUNT).value
        )

        if amount is None:
            continue

        raw_keyword = worksheet.cell(
            row,
            S1_KEYWORD,
        ).value

        match_type = worksheet.cell(
            row,
            S1_MATCH_TYPE,
        ).value

        records.append(
            {
                "row": row,
                "amount": amount,
                "amount_cents": amount_to_cents(amount),
                "keywords": keyword_set(raw_keyword),
                "keyword_raw": normalize_text(raw_keyword),
                "match_type": normalize_text(match_type),
                "unmatched": is_blank(match_type),
                "text": normalize_text(
                    worksheet.cell(row, S1_TEXT).value
                ),
            }
        )

    return records


def load_sheet2(worksheet) -> list[dict]:
    records = []

    for row in range(START_ROW, worksheet.max_row + 1):
        amount = parse_amount(
            worksheet.cell(row, S2_AMOUNT).value
        )

        if amount is None:
            continue

        raw_keyword = worksheet.cell(
            row,
            S2_KEYWORD,
        ).value

        match_type = worksheet.cell(
            row,
            S2_MATCH_TYPE,
        ).value

        records.append(
            {
                "row": row,
                "amount": amount,
                "amount_cents": amount_to_cents(amount),
                "keywords": keyword_set(raw_keyword),
                "keyword_raw": normalize_text(raw_keyword),
                "match_type": normalize_text(match_type),
                "unmatched": is_blank(match_type),
                "text": normalize_text(
                    worksheet.cell(row, S2_NAME).value
                ),
            }
        )

    return records


def sum_amounts(records: list[dict]) -> Decimal:
    return sum(
        (record["amount"] for record in records),
        Decimal("0"),
    )


def search_exact_combination(
    records: list[dict],
    target: Decimal,
    max_size: int,
    time_limit_seconds: int,
) -> tuple[list[dict] | None, str, float]:
    """
    动态规划搜索最少笔数组合。

    states[size] 保存：
        金额合计（整数分） -> 对应记录索引组合

    每条记录最多使用一次。
    """
    started = monotonic()
    target_cents = amount_to_cents(target)

    if target_cents == 0:
        return [], "TARGET_ZERO", 0.0

    # 先检查单笔。
    for index, record in enumerate(records):
        if record["amount_cents"] == target_cents:
            return (
                [record],
                "FOUND",
                monotonic() - started,
            )

    states: list[dict[int, tuple[int, ...]]] = [
        {}
        for _ in range(max_size + 1)
    ]
    states[0][0] = ()

    for index, record in enumerate(records):
        if monotonic() - started > time_limit_seconds:
            return None, "TIME_LIMIT", monotonic() - started

        value = record["amount_cents"]
        upper_size = min(max_size, index + 1)

        # 倒序更新，保证同一条记录不会重复使用。
        for size in range(upper_size, 0, -1):
            previous_states = list(
                states[size - 1].items()
            )

            for previous_sum, previous_indexes in previous_states:
                new_sum = previous_sum + value

                if new_sum in states[size]:
                    continue

                new_indexes = previous_indexes + (index,)
                states[size][new_sum] = new_indexes

                if new_sum == target_cents:
                    group = [
                        records[item_index]
                        for item_index in new_indexes
                    ]
                    return (
                        group,
                        "FOUND",
                        monotonic() - started,
                    )

                if len(states[size]) >= MAX_STATES_PER_SIZE:
                    break

            if monotonic() - started > time_limit_seconds:
                return None, "TIME_LIMIT", monotonic() - started

        if (index + 1) % 20 == 0:
            state_counts = ", ".join(
                f"{size}:{len(states[size])}"
                for size in range(1, max_size + 1)
                if states[size]
            )
            print(
                f"Processed {index + 1}/{len(records)} candidates "
                f"| states {state_counts}"
            )

    return None, "NOT_FOUND", monotonic() - started


def print_records(
    title: str,
    records: list[dict],
    source_sheet: str,
) -> None:
    print()
    print(title)
    print("-" * 72)

    if not records:
        print("(none)")
        return

    for record in records:
        print(
            f"{source_sheet} row {record['row']:<5} "
            f"Amount {str(record['amount']):>12} | "
            f"Key word: {record['keyword_raw'] or '(blank)'} | "
            f"Text: {record['text']}"
        )


def main() -> None:
    if not RESULT_FILE.exists():
        raise FileNotFoundError(
            f"找不到 {RESULT_FILE}。\n"
            "请先运行 python main.py 生成结果文件。"
        )

    workbook = load_workbook(
        RESULT_FILE,
        data_only=True,
    )

    for sheet_name in ("Sheet1", "Sheet2"):
        if sheet_name not in workbook.sheetnames:
            raise ValueError(
                f"{RESULT_FILE} 中找不到 {sheet_name}。"
            )

    sheet1 = load_sheet1(
        workbook["Sheet1"]
    )
    sheet2 = load_sheet2(
        workbook["Sheet2"]
    )

    # Sheet2 中尚未匹配的 NATSTEEL。
    sheet2_unmatched_natsteel = [
        record
        for record in sheet2
        if (
            record["unmatched"]
            and TARGET_KEYWORD in record["keywords"]
        )
    ]

    # Sheet1 中尚未匹配且已经识别为 NATSTEEL 的记录。
    sheet1_unmatched_natsteel = [
        record
        for record in sheet1
        if (
            record["unmatched"]
            and TARGET_KEYWORD in record["keywords"]
        )
    ]

    sheet2_target_total = sum_amounts(
        sheet2_unmatched_natsteel
    )
    sheet1_known_total = sum_amounts(
        sheet1_unmatched_natsteel
    )

    difference = (
        sheet2_target_total
        - sheet1_known_total
    )

    # 其余黄色 Sheet1 记录：用于寻找漏掉的 NATSTEEL 金额。
    candidates = [
        record
        for record in sheet1
        if (
            record["unmatched"]
            and TARGET_KEYWORD not in record["keywords"]
        )
    ]

    print("=" * 72)
    print("NATSTEEL Remaining One-to-Many Analysis")
    print("=" * 72)
    print(f"Source file                       : {RESULT_FILE}")
    print(
        f"Sheet2 unmatched NATSTEEL rows    : "
        f"{len(sheet2_unmatched_natsteel)}"
    )
    print(
        f"Sheet2 unmatched NATSTEEL total   : "
        f"{sheet2_target_total}"
    )
    print(
        f"Sheet1 unmatched NATSTEEL rows    : "
        f"{len(sheet1_unmatched_natsteel)}"
    )
    print(
        f"Sheet1 unmatched NATSTEEL total   : "
        f"{sheet1_known_total}"
    )
    print(
        f"Missing amount to search          : "
        f"{difference}"
    )
    print(
        f"Other unmatched Sheet1 candidates : "
        f"{len(candidates)}"
    )

    print_records(
        title="Sheet2 unmatched NATSTEEL",
        records=sheet2_unmatched_natsteel,
        source_sheet="Sheet2",
    )

    print_records(
        title="Sheet1 unmatched NATSTEEL already identified",
        records=sheet1_unmatched_natsteel,
        source_sheet="Sheet1",
    )

    print()
    print("=" * 72)
    print("Searching remaining yellow Sheet1 rows")
    print("=" * 72)

    group, status, elapsed = search_exact_combination(
        records=candidates,
        target=difference,
        max_size=MAX_COMBINATION_SIZE,
        time_limit_seconds=SEARCH_TIME_LIMIT_SECONDS,
    )

    print()
    print("=" * 72)
    print("Candidate Combination Result")
    print("=" * 72)
    print(f"Status               : {status}")
    print(f"Target difference    : {difference}")
    print(f"Search time          : {elapsed:.3f}s")

    if group is None:
        print(
            "No exact combination was found "
            "within the current search limit."
        )

    elif not group:
        print(
            "The unmatched Sheet1 NATSTEEL rows already "
            "equal the unmatched Sheet2 NATSTEEL total."
        )

    else:
        total = sum_amounts(group)

        print(
            f"Candidate item count : "
            f"{len(group)}"
        )
        print(
            f"Candidate total      : "
            f"{total}"
        )

        print_records(
            title="Additional Sheet1 candidate rows",
            records=group,
            source_sheet="Sheet1",
        )

        final_total = (
            sheet1_known_total
            + total
        )

        print()
        print(
            f"Sheet1 known NATSTEEL total : "
            f"{sheet1_known_total}"
        )
        print(
            f"Additional candidate total  : "
            f"{total}"
        )
        print(
            f"Combined Sheet1 total       : "
            f"{final_total}"
        )
        print(
            f"Sheet2 unmatched total      : "
            f"{sheet2_target_total}"
        )
        print(
            f"Final difference            : "
            f"{sheet2_target_total - final_total}"
        )

    print("=" * 72)


if __name__ == "__main__":
    main()