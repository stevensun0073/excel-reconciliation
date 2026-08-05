from collections import defaultdict
from decimal import Decimal, InvalidOperation
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill


INPUT_FILE = Path("remaining_matching_result.xlsx")
OUTPUT_FILE = Path("remaining_keyword_combination_result.xlsx")

HEADER_ROW = 2
DATA_START_ROW = 3
MAX_COMBINATION_SIZE = 6


# 本轮组合匹配成功：淡黄色
LIGHT_YELLOW_FILL = PatternFill(
    start_color="FFF2CC",
    end_color="FFF2CC",
    fill_type="solid",
)

# 最终仍未匹配：深黄色
DARK_YELLOW_FILL = PatternFill(
    start_color="FFD966",
    end_color="FFD966",
    fill_type="solid",
)


def normalize(value):
    if value is None:
        return ""

    return str(value).strip().lower()


def to_decimal(value):
    if value in (None, ""):
        return None

    if isinstance(value, bool):
        return None

    text = str(value).strip().replace(",", "")

    if text.startswith("(") and text.endswith(")"):
        text = "-" + text[1:-1]

    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def to_cents(value):
    amount = to_decimal(value)

    if amount is None:
        return None

    return int(amount * 100)


def find_column(ws, possible_headers):
    for cell in ws[HEADER_ROW]:
        if normalize(cell.value) in possible_headers:
            return cell.column

    headers = [
        str(cell.value).strip()
        for cell in ws[HEADER_ROW]
        if cell.value is not None
    ]

    raise ValueError(
        f"{ws.title} 找不到列：{sorted(possible_headers)}\n"
        f"当前表头：{headers}"
    )


def color_row(ws, row_number, fill):
    for column_number in range(1, ws.max_column + 1):
        ws.cell(row_number, column_number).fill = fill


def collect_unmatched_records(
    ws,
    original_row_col,
    amount_col,
    keyword_col,
    match_col,
    business_type_col=None,
):
    """
    只收集当前仍未匹配的记录。

    原有淡蓝色记录已经有 Match Type，
    因而不会进入后续组合计算。
    """
    records = []

    for row in range(DATA_START_ROW, ws.max_row + 1):
        match_type = normalize(
            ws.cell(row, match_col).value
        )

        if match_type:
            continue

        cents = to_cents(
            ws.cell(row, amount_col).value
        )

        if cents is None:
            continue

        business_type = ""

        if business_type_col is not None:
            business_type = normalize(
                ws.cell(row, business_type_col).value
            )

        records.append(
            {
                "row": row,
                "original_row": ws.cell(
                    row,
                    original_row_col,
                ).value,
                "cents": cents,
                "keyword": normalize(
                    ws.cell(row, keyword_col).value
                ),
                "business_type": business_type,
            }
        )

    return records


def sheet2_candidates_for_sheet1(
    sheet1_record,
    sheet2_records,
):
    """
    业务规则：

    当 Sheet1 的 Key word 是 BANK 时，
    只能匹配 Sheet2 中 Business Type 为 Charging 的记录。

    Sheet1 其他 Key word 不受此规则限制。
    """
    if sheet1_record["keyword"] == "bank":
        return [
            record
            for record in sheet2_records
            if record["business_type"] == "charging"
        ]

    return sheet2_records


def sheet1_candidates_for_sheet2(
    sheet2_record,
    sheet1_records,
):
    """
    反向应用同一业务规则：

    当 Sheet2 的 Business Type 不是 Charging 时，
    Sheet1 中 Key word 为 BANK 的记录不得参加组合。

    当 Sheet2 是 Charging 时，
    BANK 和其他 Key word 均可参加。
    """
    if sheet2_record["business_type"] != "charging":
        return [
            record
            for record in sheet1_records
            if record["keyword"] != "bank"
        ]

    return sheet1_records


def find_first_combination(
    records,
    target_cents,
    combination_size,
    used_rows,
):
    """
    寻找第一个指定条数、金额完全相等的组合。

    规则：
    1. 只使用尚未匹配的记录；
    2. 组合金额必须与目标金额同号；
    3. 金额以分为单位完全相等；
    4. 找到第一个可用组合后立即返回；
    5. 不继续搜索其他组合。
    """
    if target_cents == 0:
        return None

    candidates = []

    for record in records:
        if record["row"] in used_rows:
            continue

        cents = record["cents"]

        if target_cents > 0:
            if cents <= 0 or cents > target_cents:
                continue
        else:
            if cents >= 0 or cents < target_cents:
                continue

        candidates.append(record)

    if len(candidates) < combination_size:
        return None

    target = abs(target_cents)

    candidates.sort(
        key=lambda record: (
            abs(record["cents"]),
            record["row"],
        )
    )

    # states[count][sum] = 第一个找到的组合
    states = [
        {}
        for _ in range(combination_size + 1)
    ]

    states[0][0] = tuple()

    for record in candidates:
        value = abs(record["cents"])

        for count in range(
            combination_size - 1,
            -1,
            -1,
        ):
            current_states = list(
                states[count].items()
            )

            for current_sum, combination in current_states:
                new_sum = current_sum + value

                if new_sum > target:
                    continue

                new_count = count + 1

                if new_sum not in states[new_count]:
                    states[new_count][new_sum] = (
                        combination + (record,)
                    )

                if (
                    new_count == combination_size
                    and new_sum == target
                ):
                    return list(
                        states[new_count][new_sum]
                    )

    return None


def write_one_to_many(
    sheet1,
    sheet2,
    sheet1_record,
    sheet2_records,
    size,
    match_prefix,
    columns,
):
    (
        s1_match_col,
        s1_partner_col,
        s1_review_col,
        s2_match_col,
        s2_partner_col,
        s2_review_col,
    ) = columns

    match_type = f"{match_prefix} 1-to-{size}"

    sheet2_original_rows = ", ".join(
        str(record["original_row"])
        for record in sheet2_records
    )

    sheet1_row = sheet1_record["row"]

    sheet1.cell(
        sheet1_row,
        s1_match_col,
    ).value = match_type

    sheet1.cell(
        sheet1_row,
        s1_partner_col,
    ).value = sheet2_original_rows

    sheet1.cell(
        sheet1_row,
        s1_review_col,
    ).value = f"{match_prefix} 1对{size}"

    color_row(
        sheet1,
        sheet1_row,
        LIGHT_YELLOW_FILL,
    )

    for record in sheet2_records:
        row = record["row"]

        sheet2.cell(
            row,
            s2_match_col,
        ).value = match_type

        sheet2.cell(
            row,
            s2_partner_col,
        ).value = sheet1_record["original_row"]

        sheet2.cell(
            row,
            s2_review_col,
        ).value = f"{match_prefix} 1对{size}"

        color_row(
            sheet2,
            row,
            LIGHT_YELLOW_FILL,
        )


def write_many_to_one(
    sheet1,
    sheet2,
    sheet1_records,
    sheet2_record,
    size,
    match_prefix,
    columns,
):
    (
        s1_match_col,
        s1_partner_col,
        s1_review_col,
        s2_match_col,
        s2_partner_col,
        s2_review_col,
    ) = columns

    match_type = f"{match_prefix} {size}-to-1"

    sheet1_original_rows = ", ".join(
        str(record["original_row"])
        for record in sheet1_records
    )

    for record in sheet1_records:
        row = record["row"]

        sheet1.cell(
            row,
            s1_match_col,
        ).value = match_type

        sheet1.cell(
            row,
            s1_partner_col,
        ).value = sheet2_record["original_row"]

        sheet1.cell(
            row,
            s1_review_col,
        ).value = f"{match_prefix} {size}对1"

        color_row(
            sheet1,
            row,
            LIGHT_YELLOW_FILL,
        )

    sheet2_row = sheet2_record["row"]

    sheet2.cell(
        sheet2_row,
        s2_match_col,
    ).value = match_type

    sheet2.cell(
        sheet2_row,
        s2_partner_col,
    ).value = sheet1_original_rows

    sheet2.cell(
        sheet2_row,
        s2_review_col,
    ).value = f"{match_prefix} {size}对1"

    color_row(
        sheet2,
        sheet2_row,
        LIGHT_YELLOW_FILL,
    )


def group_by_keyword(records, used_rows):
    """
    将同一张表内仍未使用的记录按 Key word 分组。

    空白 Key word 不作为关键词优先组合。
    """
    groups = defaultdict(list)

    for record in records:
        if record["row"] in used_rows:
            continue

        keyword = record["keyword"]

        if not keyword:
            continue

        groups[keyword].append(record)

    return groups


def run_keyword_priority_matching(
    sheet1,
    sheet2,
    sheet1_records,
    sheet2_records,
    used_sheet1_rows,
    used_sheet2_rows,
    columns,
):
    """
    第一轮：同一张表内 Key word 相同的记录优先组合。

    Sheet1：
    同 Key word 的2至6条记录，匹配Sheet2一条。

    Sheet2：
    同 Key word 的2至6条记录，匹配Sheet1一条。

    两张表之间的 Key word 不要求相同。

    BANK 业务规则始终有效：
    Sheet1 BANK 只能匹配 Sheet2 Charging。
    """
    one_to_many_counts = defaultdict(int)
    many_to_one_counts = defaultdict(int)

    for size in range(2, MAX_COMBINATION_SIZE + 1):

        # ====================================================
        # Sheet1 同 Key word 多条 → Sheet2 一条
        # ====================================================

        sheet1_keyword_groups = group_by_keyword(
            sheet1_records,
            used_sheet1_rows,
        )

        for sheet2_record in sheet2_records:
            if sheet2_record["row"] in used_sheet2_rows:
                continue

            found_combination = None

            allowed_sheet1_records = (
                sheet1_candidates_for_sheet2(
                    sheet2_record,
                    sheet1_records,
                )
            )

            allowed_sheet1_rows = {
                record["row"]
                for record in allowed_sheet1_records
            }

            for keyword in sorted(sheet1_keyword_groups):
                keyword_records = [
                    record
                    for record in sheet1_keyword_groups[keyword]
                    if record["row"] in allowed_sheet1_rows
                ]

                if len(keyword_records) < size:
                    continue

                combination = find_first_combination(
                    records=keyword_records,
                    target_cents=sheet2_record["cents"],
                    combination_size=size,
                    used_rows=used_sheet1_rows,
                )

                if combination is not None:
                    found_combination = combination
                    break

            if found_combination is None:
                continue

            write_many_to_one(
                sheet1=sheet1,
                sheet2=sheet2,
                sheet1_records=found_combination,
                sheet2_record=sheet2_record,
                size=size,
                match_prefix="Keyword Group",
                columns=columns,
            )

            for record in found_combination:
                used_sheet1_rows.add(record["row"])

            used_sheet2_rows.add(sheet2_record["row"])

            many_to_one_counts[size] += 1

        # ====================================================
        # Sheet1 一条 → Sheet2 同 Key word 多条
        # ====================================================

        sheet2_keyword_groups = group_by_keyword(
            sheet2_records,
            used_sheet2_rows,
        )

        for sheet1_record in sheet1_records:
            if sheet1_record["row"] in used_sheet1_rows:
                continue

            found_combination = None

            allowed_sheet2_records = (
                sheet2_candidates_for_sheet1(
                    sheet1_record,
                    sheet2_records,
                )
            )

            allowed_sheet2_rows = {
                record["row"]
                for record in allowed_sheet2_records
            }

            for keyword in sorted(sheet2_keyword_groups):
                keyword_records = [
                    record
                    for record in sheet2_keyword_groups[keyword]
                    if record["row"] in allowed_sheet2_rows
                ]

                if len(keyword_records) < size:
                    continue

                combination = find_first_combination(
                    records=keyword_records,
                    target_cents=sheet1_record["cents"],
                    combination_size=size,
                    used_rows=used_sheet2_rows,
                )

                if combination is not None:
                    found_combination = combination
                    break

            if found_combination is None:
                continue

            write_one_to_many(
                sheet1=sheet1,
                sheet2=sheet2,
                sheet1_record=sheet1_record,
                sheet2_records=found_combination,
                size=size,
                match_prefix="Keyword Group",
                columns=columns,
            )

            used_sheet1_rows.add(sheet1_record["row"])

            for record in found_combination:
                used_sheet2_rows.add(record["row"])

            one_to_many_counts[size] += 1

    return one_to_many_counts, many_to_one_counts


def run_amount_only_matching(
    sheet1,
    sheet2,
    sheet1_records,
    sheet2_records,
    used_sheet1_rows,
    used_sheet2_rows,
    columns,
):
    """
    第二轮：对第一轮后仍未匹配的记录进行纯金额组合。

    BANK 业务规则仍然有效。
    """
    one_to_many_counts = defaultdict(int)
    many_to_one_counts = defaultdict(int)

    for size in range(2, MAX_COMBINATION_SIZE + 1):

        # ====================================================
        # Sheet1 一条 → Sheet2 多条
        # ====================================================

        for sheet1_record in sheet1_records:
            if sheet1_record["row"] in used_sheet1_rows:
                continue

            allowed_sheet2_records = (
                sheet2_candidates_for_sheet1(
                    sheet1_record,
                    sheet2_records,
                )
            )

            combination = find_first_combination(
                records=allowed_sheet2_records,
                target_cents=sheet1_record["cents"],
                combination_size=size,
                used_rows=used_sheet2_rows,
            )

            if combination is None:
                continue

            write_one_to_many(
                sheet1=sheet1,
                sheet2=sheet2,
                sheet1_record=sheet1_record,
                sheet2_records=combination,
                size=size,
                match_prefix="Amount Only",
                columns=columns,
            )

            used_sheet1_rows.add(sheet1_record["row"])

            for record in combination:
                used_sheet2_rows.add(record["row"])

            one_to_many_counts[size] += 1

        # ====================================================
        # Sheet1 多条 → Sheet2 一条
        # ====================================================

        for sheet2_record in sheet2_records:
            if sheet2_record["row"] in used_sheet2_rows:
                continue

            allowed_sheet1_records = (
                sheet1_candidates_for_sheet2(
                    sheet2_record,
                    sheet1_records,
                )
            )

            combination = find_first_combination(
                records=allowed_sheet1_records,
                target_cents=sheet2_record["cents"],
                combination_size=size,
                used_rows=used_sheet1_rows,
            )

            if combination is None:
                continue

            write_many_to_one(
                sheet1=sheet1,
                sheet2=sheet2,
                sheet1_records=combination,
                sheet2_record=sheet2_record,
                size=size,
                match_prefix="Amount Only",
                columns=columns,
            )

            for record in combination:
                used_sheet1_rows.add(record["row"])

            used_sheet2_rows.add(sheet2_record["row"])

            many_to_one_counts[size] += 1

    return one_to_many_counts, many_to_one_counts


def main():
    if not INPUT_FILE.exists():
        raise FileNotFoundError(
            f"找不到文件：{INPUT_FILE.resolve()}"
        )

    workbook = openpyxl.load_workbook(INPUT_FILE)

    sheet1 = workbook["Sheet1"]
    sheet2 = workbook["Sheet2"]

    # ========================================================
    # Sheet1列
    # ========================================================

    s1_original_col = find_column(
        sheet1,
        {"original row"},
    )

    s1_amount_col = find_column(
        sheet1,
        {"以公司代码货币计算的金额"},
    )

    s1_keyword_col = find_column(
        sheet1,
        {"key word", "keyword"},
    )

    s1_match_col = find_column(
        sheet1,
        {"match type", "匹配类型"},
    )

    s1_partner_col = find_column(
        sheet1,
        {"partner rows", "对方行号"},
    )

    s1_review_col = find_column(
        sheet1,
        {"review", "复核"},
    )

    # ========================================================
    # Sheet2列
    # ========================================================

    s2_original_col = find_column(
        sheet2,
        {"original row"},
    )

    s2_amount_col = find_column(
        sheet2,
        {"transaction amount"},
    )

    s2_keyword_col = find_column(
        sheet2,
        {"key word", "keyword"},
    )

    s2_business_type_col = find_column(
        sheet2,
        {"business type"},
    )

    s2_match_col = find_column(
        sheet2,
        {"match type", "匹配类型"},
    )

    s2_partner_col = find_column(
        sheet2,
        {"partner rows", "对方行号"},
    )

    s2_review_col = find_column(
        sheet2,
        {"review", "复核"},
    )

    columns = (
        s1_match_col,
        s1_partner_col,
        s1_review_col,
        s2_match_col,
        s2_partner_col,
        s2_review_col,
    )

    # ========================================================
    # 只读取淡蓝色规则之后仍未匹配的记录
    # ========================================================

    sheet1_records = collect_unmatched_records(
        ws=sheet1,
        original_row_col=s1_original_col,
        amount_col=s1_amount_col,
        keyword_col=s1_keyword_col,
        match_col=s1_match_col,
    )

    sheet2_records = collect_unmatched_records(
        ws=sheet2,
        original_row_col=s2_original_col,
        amount_col=s2_amount_col,
        keyword_col=s2_keyword_col,
        match_col=s2_match_col,
        business_type_col=s2_business_type_col,
    )

    # 当前未匹配记录先标为深黄色。
    for record in sheet1_records:
        color_row(
            sheet1,
            record["row"],
            DARK_YELLOW_FILL,
        )

    for record in sheet2_records:
        color_row(
            sheet2,
            record["row"],
            DARK_YELLOW_FILL,
        )

    used_sheet1_rows = set()
    used_sheet2_rows = set()

    # ========================================================
    # 第一轮：相同 Key word 组优先
    # ========================================================

    (
        keyword_one_to_many,
        keyword_many_to_one,
    ) = run_keyword_priority_matching(
        sheet1=sheet1,
        sheet2=sheet2,
        sheet1_records=sheet1_records,
        sheet2_records=sheet2_records,
        used_sheet1_rows=used_sheet1_rows,
        used_sheet2_rows=used_sheet2_rows,
        columns=columns,
    )

    # ========================================================
    # 第二轮：剩余记录纯金额组合
    # ========================================================

    (
        amount_one_to_many,
        amount_many_to_one,
    ) = run_amount_only_matching(
        sheet1=sheet1,
        sheet2=sheet2,
        sheet1_records=sheet1_records,
        sheet2_records=sheet2_records,
        used_sheet1_rows=used_sheet1_rows,
        used_sheet2_rows=used_sheet2_rows,
        columns=columns,
    )

    workbook.save(OUTPUT_FILE)

    remaining_sheet1 = (
        len(sheet1_records) - len(used_sheet1_rows)
    )

    remaining_sheet2 = (
        len(sheet2_records) - len(used_sheet2_rows)
    )

    print("=" * 76)
    print("Keyword Priority Combination Matching")
    print("=" * 76)

    print()
    print("Business Rule")
    print("  Sheet1 Key word BANK")
    print("  can only match Sheet2 Business Type Charging")

    print()
    print("Round 1 - Same Key Word Groups")

    for size in range(2, MAX_COMBINATION_SIZE + 1):
        print(
            f"  Keyword 1-to-{size} : "
            f"{keyword_one_to_many[size]}"
        )
        print(
            f"  Keyword {size}-to-1 : "
            f"{keyword_many_to_one[size]}"
        )

    print()
    print("Round 2 - Amount Only")

    for size in range(2, MAX_COMBINATION_SIZE + 1):
        print(
            f"  Amount 1-to-{size}  : "
            f"{amount_one_to_many[size]}"
        )
        print(
            f"  Amount {size}-to-1  : "
            f"{amount_many_to_one[size]}"
        )

    print()
    print("Result")
    print(
        f"  Newly matched Sheet1 rows : "
        f"{len(used_sheet1_rows)}"
    )
    print(
        f"  Newly matched Sheet2 rows : "
        f"{len(used_sheet2_rows)}"
    )
    print(
        f"  Remaining Sheet1          : "
        f"{remaining_sheet1}"
    )
    print(
        f"  Remaining Sheet2          : "
        f"{remaining_sheet2}"
    )

    print()
    print(f"Saved: {OUTPUT_FILE.resolve()}")


if __name__ == "__main__":
    main()