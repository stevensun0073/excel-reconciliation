import itertools
from decimal import Decimal
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ============================================================
# CONFIG
# ============================================================

INPUT_FILE = "data.xlsx"
OUTPUT_FILE = "result.xlsx"

SHEET1_INDEX = 0
SHEET2_INDEX = 1
START_ROW = 3

TOLERANCE = Decimal("0.01")

# ============================================================
# COLORS
# ============================================================

COLORS = {
    "EXACT": PatternFill(fill_type="solid", fgColor="C6EFCE"),
    "TOLERANCE": PatternFill(fill_type="solid", fgColor="BDD7EE"),
    "ONE_TWO": PatternFill(fill_type="solid", fgColor="FFF2CC"),
    "ONE_THREE": PatternFill(fill_type="solid", fgColor="F4B183"),
    "TWO_TWO": PatternFill(fill_type="solid", fgColor="D9D2E9"),
    "UNMATCHED": PatternFill(fill_type="solid", fgColor="F4CCCC"),
}

# ============================================================
# RECORD
# ============================================================

class Record:

    def __init__(self, worksheet, row, cell):
        self.ws = worksheet
        self.row = row
        self.cell = cell

        self.value = Decimal(str(cell.value))

        self.matched = False
        self.match_type = ""

    def mark(self, match_type):
        self.matched = True
        self.match_type = match_type


# ============================================================
# LOAD EXCEL
# ============================================================

def load_sheet(ws):

    records = []

    for row in range(START_ROW, ws.max_row + 1):

        cell = ws.cell(row, 1)

        if cell.value is None:
            continue

        try:
            Decimal(str(cell.value))
        except Exception:
            continue

        records.append(
            Record(
                ws,
                row,
                cell
            )
        )

    return records


print("Loading Excel...")

workbook = load_workbook(INPUT_FILE)

sheet_left = workbook.worksheets[SHEET1_INDEX]
sheet_right = workbook.worksheets[SHEET2_INDEX]

left_records = load_sheet(sheet_left)
right_records = load_sheet(sheet_right)

print(f"Sheet1 : {len(left_records)} rows")
print(f"Sheet2 : {len(right_records)} rows")

# ============================================================
# COMMON FUNCTIONS
# ============================================================

def equal(a, b):
    return abs(a - b) <= TOLERANCE


def unmatched(records):
    return [r for r in records if not r.matched]


def paint(records, style):
    fill = COLORS[style]

    for r in records:
        r.cell.fill = fill


def mark(left, right, style):

    for r in left:
        r.mark(style)

    for r in right:
        r.mark(style)

    paint(left, style)
    paint(right, style)


def count(records, style):

    total = 0

    for r in records:
        if r.match_type == style:
            total += 1

    return total


print("Initialization completed.")

# ============================================================
# RECONCILIATION ENGINE
# ============================================================

class ReconciliationEngine:

    def __init__(self, left_records, right_records):

        self.left = left_records
        self.right = right_records

        self.statistics = {
            "EXACT": 0,
            "TOLERANCE": 0,
            "ONE_TWO": 0,
            "ONE_THREE": 0,
            "TWO_TWO": 0
        }

    # --------------------------------------------------------
    # Build value index
    # --------------------------------------------------------

    def build_index(self):

        index = {}

        for r in unmatched(self.right):

            index.setdefault(r.value, []).append(r)

        return index

    # --------------------------------------------------------
    # 1 ↔ 1 Exact Match
    # --------------------------------------------------------

    def exact_match(self):

        print("Running Exact Match...")

        index = self.build_index()

        for left in unmatched(self.left):

            if left.value not in index:
                continue

            if len(index[left.value]) == 0:
                continue

            right = index[left.value].pop()

            mark([left], [right], "EXACT")

            self.statistics["EXACT"] += 1

        print(f"Exact : {self.statistics['EXACT']}")

    # --------------------------------------------------------
    # 1 ↔ 1 Tolerance Match
    # --------------------------------------------------------

    def tolerance_match(self):

        print("Running Tolerance Match...")

        for left in unmatched(self.left):

            found = False

            for right in unmatched(self.right):

                if equal(left.value, right.value):

                    mark([left], [right], "TOLERANCE")

                    self.statistics["TOLERANCE"] += 1

                    found = True

                    break

            if found:
                continue

        print(f"Tolerance : {self.statistics['TOLERANCE']}")

    # --------------------------------------------------------
    # Run Step 1
    # --------------------------------------------------------

    def run_stage_one(self):

        self.exact_match()

        self.tolerance_match()