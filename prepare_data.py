"""
prepare_data.py

独立的数据准备程序，在 main.py 之前运行。

固定文件格式
------------
Sheet1：
A  银行流水号
B  文本
C  以公司代码货币计算的金额
D  Key word
E  Match Type
F  Partner Rows
G  Review

Sheet2：
A  Transaction Time
B  Recipient's Account Name
C  Business Type
D  Description
E  Transaction Amount
F  Key word
G  Match Type
H  Partner Rows
I  Review

功能
----
1. Sheet1
   - 读取“文本”列。
   - 读取 company.xlsx 已整理好的“Company Keyword”列。
   - 按完整单词或完整组合进行匹配。
   - Key word 永远写入 D 列。
   - 一条文本可以输出多个关键词，用“; ”分隔。
   - KZ、GR 使用特殊业务规则。

2. Sheet2
   - 不与 company.xlsx 匹配。
   - 直接读取“Recipient's Account Name”。
   - 按公司名称提取规则生成一个 Key word。
   - Key word 永远写入 F 列。

3. 固定格式
   - 不再在最右侧新增 Key word 列。
   - 不删除列。
   - 不移动原始数据。
   - 不修改第一行公式。
   - 会清理固定列以外残留的旧 Key word 列内容。

本程序不会执行金额对账，也不会自动运行 main.py。
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Pattern

from openpyxl import load_workbook
from openpyxl.styles import Font


COMPANY_FILE = Path("company.xlsx")
DATA_FILE = Path("data.xlsx")

COMPANY_KEYWORD_HEADER = "Company Keyword"

SHEET1_TEXT_HEADER = "文本"
SHEET2_NAME_HEADER = "Recipient's Account Name"

OUTPUT_HEADER = "Key word"

HEADER_SEARCH_ROWS = 5

# 固定输出列
SHEET1_KEYWORD_COLUMN = 4  # D列
SHEET2_KEYWORD_COLUMN = 6  # F列


# ============================================================
# 通用工具
# ============================================================

def normalize_spaces(value: str) -> str:
    """把连续空格合并为一个空格，并删除首尾空格。"""

    return re.sub(r"\s+", " ", value).strip()


def require_file(path: Path) -> None:
    """确认文件存在。"""

    if not path.exists():
        raise FileNotFoundError(
            f"找不到文件：{path}\n"
            "请确认文件位于当前项目目录，"
            "并检查文件名大小写。"
        )


def find_header(
    worksheet,
    header_name: str,
    search_rows: int = HEADER_SEARCH_ROWS,
) -> tuple[int, int]:
    """
    查找指定表头。

    返回：
        (表头行号, 列号)
    """

    last_row = min(
        search_rows,
        worksheet.max_row,
    )

    for row in range(1, last_row + 1):
        for column in range(
            1,
            worksheet.max_column + 1,
        ):
            value = worksheet.cell(
                row=row,
                column=column,
            ).value

            if value is None:
                continue

            if (
                str(value).strip().casefold()
                == header_name.casefold()
            ):
                return row, column

    raise ValueError(
        f'在工作表“{worksheet.title}”'
        f'前 {last_row} 行中找不到表头'
        f'“{header_name}”。'
    )


def prepare_fixed_output_column(
    worksheet,
    header_row: int,
    output_column: int,
    header_name: str,
) -> None:
    """
    设置固定的 Key word 输出列。

    同时清理其他列中残留的旧 Key word 表头及其数据，
    但不会删除列，也不会移动任何数据。
    """

    # 清理固定列之外重复的旧 Key word 列。
    for column in range(
        1,
        worksheet.max_column + 1,
    ):
        if column == output_column:
            continue

        header_value = worksheet.cell(
            row=header_row,
            column=column,
        ).value

        if header_value is None:
            continue

        if (
            str(header_value).strip().casefold()
            != header_name.casefold()
        ):
            continue

        worksheet.cell(
            row=header_row,
            column=column,
        ).value = None

        for row in range(
            header_row + 1,
            worksheet.max_row + 1,
        ):
            worksheet.cell(
                row=row,
                column=column,
            ).value = None

    # 固定位置写入表头。
    header_cell = worksheet.cell(
        row=header_row,
        column=output_column,
    )

    header_cell.value = header_name
    header_cell.font = Font(bold=True)

    # 清除固定列的旧关键词结果。
    for row in range(
        header_row + 1,
        worksheet.max_row + 1,
    ):
        worksheet.cell(
            row=row,
            column=output_column,
        ).value = None


def unique_in_order(values: list[str]) -> list[str]:
    """去重，同时保留原来的顺序。"""

    result: list[str] = []
    seen: set[str] = set()

    for value in values:
        key = value.casefold()

        if key in seen:
            continue

        seen.add(key)
        result.append(value)

    return result


# ============================================================
# KZ / GR 特殊匹配规则
# ============================================================

KZ_PATTERN = re.compile(
    r"(?<![A-Za-z0-9])"
    r"(?:KZ|KASZON)"
    r"(?![A-Za-z0-9])",
    re.IGNORECASE,
)

GR_DIRECT_PATTERN = re.compile(
    r"(?<![A-Za-z0-9])"
    r"(?:GR|GREAT\s+RESOURCES)"
    r"(?![A-Za-z0-9])",
    re.IGNORECASE,
)

GR_SUFFIX_PATTERN = re.compile(
    r"(?<![A-Za-z0-9])"
    r"[A-Za-z0-9]+"
    r"(?:\s+[A-Za-z0-9]+)*"
    r"\s*-\s*"
    r"(?:EL|ACMV|FP|PS)"
    r"(?![A-Za-z0-9])",
    re.IGNORECASE,
)


def match_kz(text: str) -> bool:
    """判断文本是否满足 KZ 特殊规则。"""

    return bool(KZ_PATTERN.search(text))


def match_gr(text: str) -> bool:
    """判断文本是否满足 GR 特殊规则。"""

    return (
        bool(GR_DIRECT_PATTERN.search(text))
        or bool(GR_SUFFIX_PATTERN.search(text))
    )


# ============================================================
# Sheet1：使用 company.xlsx 的 Company Keyword 匹配
# ============================================================

def contains_chinese(text: str) -> bool:
    """判断文本是否包含中文字符。"""

    return bool(
        re.search(r"[\u4e00-\u9fff]", text)
    )


def build_normal_keyword_pattern(
    keyword: str,
) -> Pattern[str]:
    """
    建立完整单词或完整组合匹配规则。

    例如：
    ARM 可以匹配 PAYMENT TO ARM，
    但不能匹配 FARM 或 ARMY。
    """

    parts = keyword.split()

    if len(parts) > 1:
        body = r"\s+".join(
            re.escape(part)
            for part in parts
        )
    else:
        body = re.escape(keyword)

    return re.compile(
        rf"(?<![A-Za-z0-9])"
        rf"{body}"
        rf"(?![A-Za-z0-9])",
        re.IGNORECASE,
    )


def load_company_keywords(
) -> list[tuple[str, Pattern[str] | None]]:
    """
    从 company.xlsx 读取唯一的 Company Keyword。

    KZ 和 GR 不放入普通关键词列表，
    因为它们使用专门的业务规则。
    """

    require_file(COMPANY_FILE)

    workbook = load_workbook(
        COMPANY_FILE,
        data_only=True,
        read_only=True,
    )

    worksheet = workbook[
        workbook.sheetnames[0]
    ]

    header_row, keyword_column = find_header(
        worksheet,
        COMPANY_KEYWORD_HEADER,
    )

    keywords: list[str] = []
    seen: set[str] = set()

    for row in range(
        header_row + 1,
        worksheet.max_row + 1,
    ):
        value = worksheet.cell(
            row=row,
            column=keyword_column,
        ).value

        if value is None:
            continue

        keyword = normalize_spaces(
            str(value)
        )

        if not keyword:
            continue

        normalized = keyword.casefold()

        if normalized in {"kz", "gr"}:
            continue

        if normalized in seen:
            continue

        seen.add(normalized)
        keywords.append(keyword)

    # 较长组合优先。
    keywords.sort(
        key=lambda item: (
            -len(item),
            item.casefold(),
        )
    )

    compiled: list[
        tuple[str, Pattern[str] | None]
    ] = []

    for keyword in keywords:
        if contains_chinese(keyword):
            compiled.append(
                (keyword, None)
            )
        else:
            compiled.append(
                (
                    keyword,
                    build_normal_keyword_pattern(
                        keyword
                    ),
                )
            )

    return compiled


def match_sheet1_text(
    text: str,
    normal_keywords: list[
        tuple[str, Pattern[str] | None]
    ],
) -> list[str]:
    """对 Sheet1 文本提取全部符合规则的关键词。"""

    matches: list[str] = []

    # 特殊规则优先。
    if match_kz(text):
        matches.append("KZ")

    if match_gr(text):
        matches.append("GR")

    # 普通 Company Keyword。
    for keyword, pattern in normal_keywords:
        if pattern is None:
            if keyword in text:
                matches.append(keyword)
        elif pattern.search(text):
            matches.append(keyword)

    return unique_in_order(matches)


def process_sheet1(
    workbook,
    normal_keywords: list[
        tuple[str, Pattern[str] | None]
    ],
) -> dict[str, int]:
    """
    处理 Sheet1。

    Key word 固定写入 D 列。
    """

    if "Sheet1" not in workbook.sheetnames:
        raise ValueError(
            'data.xlsx 中找不到工作表“Sheet1”。'
        )

    worksheet = workbook["Sheet1"]

    header_row, text_column = find_header(
        worksheet,
        SHEET1_TEXT_HEADER,
    )

    prepare_fixed_output_column(
        worksheet=worksheet,
        header_row=header_row,
        output_column=SHEET1_KEYWORD_COLUMN,
        header_name=OUTPUT_HEADER,
    )

    total_rows = (
        worksheet.max_row - header_row
    )

    matched_rows = 0
    unmatched_rows = 0
    empty_rows = 0
    multiple_rows = 0

    print(
        f"Processing Sheet1: "
        f"{total_rows} rows..."
    )

    for row in range(
        header_row + 1,
        worksheet.max_row + 1,
    ):
        processed = row - header_row

        if processed % 200 == 0:
            print(
                f"  Sheet1 processed "
                f"{processed}/{total_rows} rows..."
            )

        raw_text = worksheet.cell(
            row=row,
            column=text_column,
        ).value

        if (
            raw_text is None
            or not str(raw_text).strip()
        ):
            empty_rows += 1
            continue

        text = normalize_spaces(
            str(raw_text)
        )

        matches = match_sheet1_text(
            text,
            normal_keywords,
        )

        if not matches:
            unmatched_rows += 1
            continue

        worksheet.cell(
            row=row,
            column=SHEET1_KEYWORD_COLUMN,
        ).value = "; ".join(matches)

        matched_rows += 1

        if len(matches) > 1:
            multiple_rows += 1

    return {
        "total": total_rows,
        "matched": matched_rows,
        "unmatched": unmatched_rows,
        "empty": empty_rows,
        "multiple": multiple_rows,
        "output_column": (
            SHEET1_KEYWORD_COLUMN
        ),
    }


# ============================================================
# Sheet2：从 Recipient's Account Name 提取关键词
# ============================================================

def clean_company_name(
    company_name: str,
) -> str:
    """删除括号说明并整理空格。"""

    text = re.sub(
        r"\([^)]*\)",
        " ",
        company_name,
    )

    return normalize_spaces(text)


def is_single_letter(token: str) -> bool:
    """判断是否为单个英文字母。"""

    return bool(
        re.fullmatch(r"[A-Za-z]", token)
    )


def is_only_number(token: str) -> bool:
    """判断是否为纯数字。"""

    return bool(
        re.fullmatch(r"\d+", token)
    )


def extract_sheet2_keyword(
    account_name: str,
) -> str:
    """
    从 Recipient's Account Name 提取 Key word。
    """

    text = normalize_spaces(
        str(account_name)
    )

    if not text:
        return ""

    # 中文规则优先。
    if "零星" in text:
        return "零星"

    if contains_chinese(text):
        return "中国"

    # KZ / GR 特殊规则优先。
    if match_kz(text):
        return "KZ"

    if match_gr(text):
        return "GR"

    cleaned = clean_company_name(text)
    tokens = cleaned.split()

    if not tokens:
        return ""

    first = tokens[0]

    # THE 或 SINGAPORE 开头时取第二个词。
    if first.casefold() in {
        "the",
        "singapore",
    }:
        if len(tokens) >= 2:
            return tokens[1]

        return first

    # A-B、A/B、PRO-WERKZE 等完整组合。
    if re.fullmatch(
        r"[A-Za-z0-9]+"
        r"(?:[-/][A-Za-z0-9]+)+",
        first,
    ):
        return first

    # 数字开头：数字加后面一个词。
    if is_only_number(first):
        if len(tokens) >= 2:
            return (
                f"{first} {tokens[1]}"
            )

        return first

    # A & B 作为整体。
    if (
        len(tokens) >= 3
        and is_single_letter(tokens[0])
        and tokens[1] == "&"
        and is_single_letter(tokens[2])
    ):
        return (
            f"{tokens[0]} & {tokens[2]}"
        )

    # 一个或多个单字母开头，
    # 与后面第一个有效词组合。
    if is_single_letter(first):
        parts: list[str] = []

        for token in tokens:
            parts.append(token)

            if token in {"&", "-", "/"}:
                continue

            if not is_single_letter(token):
                break

        return normalize_spaces(
            " ".join(parts)
        )

    # 普通名称取第一个词。
    return first


def process_sheet2(
    workbook,
) -> dict[str, int]:
    """
    处理 Sheet2。

    Key word 固定写入 F 列。
    """

    if "Sheet2" not in workbook.sheetnames:
        raise ValueError(
            'data.xlsx 中找不到工作表“Sheet2”。'
        )

    worksheet = workbook["Sheet2"]

    header_row, name_column = find_header(
        worksheet,
        SHEET2_NAME_HEADER,
    )

    prepare_fixed_output_column(
        worksheet=worksheet,
        header_row=header_row,
        output_column=SHEET2_KEYWORD_COLUMN,
        header_name=OUTPUT_HEADER,
    )

    total_rows = (
        worksheet.max_row - header_row
    )

    generated_rows = 0
    empty_rows = 0

    print(
        f"Processing Sheet2: "
        f"{total_rows} rows..."
    )

    for row in range(
        header_row + 1,
        worksheet.max_row + 1,
    ):
        processed = row - header_row

        if processed % 200 == 0:
            print(
                f"  Sheet2 processed "
                f"{processed}/{total_rows} rows..."
            )

        raw_name = worksheet.cell(
            row=row,
            column=name_column,
        ).value

        if (
            raw_name is None
            or not str(raw_name).strip()
        ):
            empty_rows += 1
            continue

        keyword = extract_sheet2_keyword(
            str(raw_name)
        )

        if keyword:
            worksheet.cell(
                row=row,
                column=SHEET2_KEYWORD_COLUMN,
            ).value = keyword

            generated_rows += 1

    return {
        "total": total_rows,
        "generated": generated_rows,
        "empty": empty_rows,
        "output_column": (
            SHEET2_KEYWORD_COLUMN
        ),
    }


# ============================================================
# 主程序
# ============================================================

def main() -> None:
    """执行数据准备。"""

    require_file(DATA_FILE)

    print(
        "Loading Company Keyword "
        "master data..."
    )

    normal_keywords = (
        load_company_keywords()
    )

    print(
        f"Loaded {len(normal_keywords)} "
        f"unique normal keywords."
    )

    print(
        "Special rules enabled: KZ, GR"
    )

    print("Opening data.xlsx...")

    workbook = load_workbook(DATA_FILE)

    sheet1_result = process_sheet1(
        workbook,
        normal_keywords,
    )

    sheet2_result = process_sheet2(
        workbook
    )

    print("Saving data.xlsx...")

    workbook.save(DATA_FILE)

    print("=" * 60)
    print("Prepare Data Finished")
    print("=" * 60)

    print("Sheet1")
    print(
        f"  Rows processed      : "
        f"{sheet1_result['total']}"
    )
    print(
        f"  Matched rows        : "
        f"{sheet1_result['matched']}"
    )
    print(
        f"  Unmatched rows      : "
        f"{sheet1_result['unmatched']}"
    )
    print(
        f"  Empty text rows     : "
        f"{sheet1_result['empty']}"
    )
    print(
        f"  Multiple matches    : "
        f"{sheet1_result['multiple']}"
    )
    print(
        f"  Key word column     : "
        f"{sheet1_result['output_column']}"
    )

    print("Sheet2")
    print(
        f"  Rows processed      : "
        f"{sheet2_result['total']}"
    )
    print(
        f"  Keywords generated  : "
        f"{sheet2_result['generated']}"
    )
    print(
        f"  Empty name rows     : "
        f"{sheet2_result['empty']}"
    )
    print(
        f"  Key word column     : "
        f"{sheet2_result['output_column']}"
    )

    print(
        f"Saved file            : "
        f"{DATA_FILE.resolve()}"
    )

    print("=" * 60)


if __name__ == "__main__":
    main()