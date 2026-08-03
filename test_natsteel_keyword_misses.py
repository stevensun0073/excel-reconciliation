"""
test_natsteel_keyword_misses.py

Standalone diagnostic test.

Purpose
-------
Find Sheet1 rows where the original Text clearly contains NATSTEEL-related
wording, but the Key word column does not contain NATSTEEL.

The script reads result_reconciliation.xlsx and prints:
- original row
- amount
- current Key word
- Match Type
- Partner Rows
- original Text

It does not modify any Excel file.
"""

from __future__ import annotations

import re
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook


RESULT_FILE = Path("result_reconciliation.xlsx")

START_ROW = 3

S1_TEXT = 2
S1_AMOUNT = 3
S1_KEYWORD = 4
S1_MATCH_TYPE = 5
S1_PARTNER_ROWS = 6

TARGET_KEYWORD = "NATSTEEL"


def normalize_text(value) -> str:
    if value is None:
        return ""

    return " ".join(str(value).strip().split())


def parse_amount(value) -> Decimal | None:
    if value is None:
        return None

    try:
        cleaned = str(value).replace(",", "").strip()

        if not cleaned:
            return None

        return Decimal(cleaned)

    except (InvalidOperation, ValueError, TypeError):
        return None


def keyword_set(value) -> set[str]:
    if value is None:
        return set()

    return {
        part.strip().upper()
        for part in str(value).split(";")
        if part.strip()
    }


def text_looks_like_natsteel(text: str) -> bool:
    """
    Conservative text test.

    Match:
    - NATSTEEL as a standalone part
    - NAT-STEEL
    - NAT STEEL

    Do not match arbitrary partial letter sequences inside longer words.
    """
    patterns = (
        r"(?<![A-Z0-9])NATSTEEL(?![A-Z0-9])",
        r"(?<![A-Z0-9])NAT[\s\-_]+STEEL(?![A-Z0-9])",
    )

    upper_text = text.upper()

    return any(
        re.search(pattern, upper_text)
        for pattern in patterns
    )


def main() -> None:
    if not RESULT_FILE.exists():
        raise FileNotFoundError(
            f"找不到 {RESULT_FILE}。请先运行 python main.py。"
        )

    workbook = load_workbook(
        RESULT_FILE,
        data_only=True,
    )

    if "Sheet1" not in workbook.sheetnames:
        raise ValueError(
            f"{RESULT_FILE} 中找不到 Sheet1。"
        )

    worksheet = workbook["Sheet1"]

    suspicious_rows = []

    for row in range(START_ROW, worksheet.max_row + 1):
        amount = parse_amount(
            worksheet.cell(row, S1_AMOUNT).value
        )

        if amount is None:
            continue

        text = normalize_text(
            worksheet.cell(row, S1_TEXT).value
        )

        if not text_looks_like_natsteel(text):
            continue

        raw_keyword = worksheet.cell(
            row,
            S1_KEYWORD,
        ).value

        keywords = keyword_set(raw_keyword)

        if TARGET_KEYWORD in keywords:
            continue

        suspicious_rows.append(
            {
                "row": row,
                "amount": amount,
                "keyword": normalize_text(raw_keyword),
                "match_type": normalize_text(
                    worksheet.cell(
                        row,
                        S1_MATCH_TYPE,
                    ).value
                ),
                "partner_rows": normalize_text(
                    worksheet.cell(
                        row,
                        S1_PARTNER_ROWS,
                    ).value
                ),
                "text": text,
            }
        )

    print("=" * 78)
    print("NATSTEEL Keyword Miss Diagnostic")
    print("=" * 78)
    print(f"Source file        : {RESULT_FILE}")
    print(
        f"Suspicious rows    : "
        f"{len(suspicious_rows)}"
    )
    print("-" * 78)

    if not suspicious_rows:
        print(
            "No Sheet1 row was found where Text contains NATSTEEL "
            "but Key word does not contain NATSTEEL."
        )
        print("=" * 78)
        return

    total = sum(
        (record["amount"] for record in suspicious_rows),
        Decimal("0"),
    )

    for record in suspicious_rows:
        print(
            f"Row {record['row']:<5} "
            f"Amount {str(record['amount']):>12} | "
            f"Key word: {record['keyword'] or '(blank)'} | "
            f"Match Type: {record['match_type'] or '(unmatched)'} | "
            f"Partner Rows: {record['partner_rows'] or '(blank)'}"
        )
        print(
            f"    Text: {record['text']}"
        )

    print("-" * 78)
    print(f"Suspicious amount total : {total}")
    print("=" * 78)


if __name__ == "__main__":
    main()