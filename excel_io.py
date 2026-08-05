from decimal import Decimal, InvalidOperation
from typing import TYPE_CHECKING

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from models import Record

if TYPE_CHECKING:
    from difference_analyzer import DifferenceAnalysisResult


HEADER_ROW = 2
START_ROW = 3


# ============================================================
# 固定文件格式
#
# Sheet1：
# A 银行流水号
# B 文本
# C 金额
# D Key word
# E Match Type
# F Partner Rows
# G Review
#
# Sheet2：
# A Transaction Time
# B Recipient's Account Name
# C Business Type
# D Description
# E Transaction Amount
# F Key word
# G Match Type
# H Partner Rows
# I Review
# ============================================================

SHEET_CONFIG = {
    "Sheet1": {
        "amount_column": 3,       # C列
        "key_word_column": 4,     # D列
        "match_type_column": 5,   # E列
        "partner_column": 6,      # F列
        "review_column": 7,       # G列
    },
    "Sheet2": {
        "amount_column": 5,       # E列
        "key_word_column": 6,     # F列
        "match_type_column": 7,   # G列
        "partner_column": 8,      # H列
        "review_column": 9,       # I列
    },
}


# 绿色：原有关键词相同匹配
MATCHED_KEYWORD_SAME_FILL = PatternFill(
    start_color="93C47D",
    end_color="93C47D",
    fill_type="solid",
)

# 棕色：原有可靠但关键词不同匹配
MATCHED_KEYWORD_DIFFERENT_FILL = PatternFill(
    start_color="E3D5CA",
    end_color="E3D5CA",
    fill_type="solid",
)

# 淡蓝色：
# 1. 双方唯一金额一对一
# 2. Sheet1同银行流水号多对一
BLUE_MATCH_FILL = PatternFill(
    start_color="BDD7EE",
    end_color="BDD7EE",
    fill_type="solid",
)

# 淡黄色：
# Key word组优先组合及后续纯金额组合
LIGHT_YELLOW_MATCH_FILL = PatternFill(
    start_color="FFF2CC",
    end_color="FFF2CC",
    fill_type="solid",
)

# 深黄色：最终仍未匹配
UNMATCHED_FILL = PatternFill(
    start_color="FFD966",
    end_color="FFD966",
    fill_type="solid",
)


def parse_amount(value):
    """把Excel金额安全转换为Decimal。"""

    if value is None:
        return None

    try:
        cleaned_value = (
            str(value)
            .replace(",", "")
            .strip()
        )

        if not cleaned_value:
            return None

        if (
            cleaned_value.startswith("(")
            and cleaned_value.endswith(")")
        ):
            cleaned_value = (
                "-"
                + cleaned_value[1:-1]
            )

        return Decimal(cleaned_value)

    except (
        InvalidOperation,
        AttributeError,
        ValueError,
    ):
        return None


def get_sheet_config(sheet_name):
    """取得指定工作表的列配置。"""

    if sheet_name not in SHEET_CONFIG:
        raise ValueError(
            f"没有为工作表 {sheet_name} 设置读取规则。"
        )

    return SHEET_CONFIG[sheet_name]


def read_headers(sheet, last_data_column):
    """
    读取第2行的原始字段名称。

    Sheet1读取A至D；
    Sheet2读取A至F。

    因此Sheet2的Business Type会保存在record.extra中，
    供BANK只能匹配Charging规则使用。
    """

    headers = {}

    for column in range(
        1,
        last_data_column + 1,
    ):
        value = sheet.cell(
            row=HEADER_ROW,
            column=column,
        ).value

        if (
            value is None
            or str(value).strip() == ""
        ):
            header = (
                f"Column "
                f"{get_column_letter(column)}"
            )
        else:
            header = str(value).strip()

        headers[column] = header

    return headers


def read_sheet(sheet):
    """
    按固定格式读取金额、Key word及原始辅助字段。
    """

    config = get_sheet_config(sheet.title)

    amount_column = config["amount_column"]
    key_word_column = config["key_word_column"]

    headers = read_headers(
        sheet=sheet,
        last_data_column=key_word_column,
    )

    records = []

    for row in range(
        START_ROW,
        sheet.max_row + 1,
    ):
        amount = parse_amount(
            sheet.cell(
                row=row,
                column=amount_column,
            ).value
        )

        if amount is None:
            continue

        extra = {}

        for column in range(
            1,
            key_word_column + 1,
        ):
            if column == amount_column:
                continue

            header = headers[column]

            value = sheet.cell(
                row=row,
                column=column,
            ).value

            extra[header] = value

        records.append(
            Record(
                row=row,
                amount=amount,
                source_sheet=sheet.title,
                extra=extra,
            )
        )

    return records


def load_excel(filename):
    """打开工作簿并读取Sheet1和Sheet2。"""

    workbook = load_workbook(filename)

    for sheet_name in (
        "Sheet1",
        "Sheet2",
    ):
        if sheet_name not in workbook.sheetnames:
            raise ValueError(
                f"找不到工作表 {sheet_name}。"
            )

    return (
        workbook,
        read_sheet(workbook["Sheet1"]),
        read_sheet(workbook["Sheet2"]),
    )


def get_fill(record):
    """
    根据最终匹配状态返回颜色。
    """

    if not record.matched:
        return UNMATCHED_FILL

    if record.match_type.startswith("Blue "):
        return BLUE_MATCH_FILL

    if record.match_type.startswith("Yellow "):
        return LIGHT_YELLOW_MATCH_FILL

    if record.keyword_conflict:
        return MATCHED_KEYWORD_DIFFERENT_FILL

    return MATCHED_KEYWORD_SAME_FILL


def get_result_last_column(sheet):
    """取得结果区域最后一列。"""

    config = get_sheet_config(sheet.title)

    return config["review_column"]


def fill_row(sheet, row, fill):
    """
    给数据行和结果列着色。

    不处理第1行，因此不会影响求和公式。
    """

    last_column = get_result_last_column(sheet)

    for column in range(
        1,
        last_column + 1,
    ):
        sheet.cell(
            row=row,
            column=column,
        ).fill = fill


def clear_old_results(sheet):
    """
    清除上次运行留下的匹配结果和颜色。

    不增加列；
    不删除列；
    不移动原始数据；
    不增加工作表；
    不修改第1行求和公式。
    """

    config = get_sheet_config(sheet.title)

    match_type_column = (
        config["match_type_column"]
    )

    partner_column = (
        config["partner_column"]
    )

    review_column = (
        config["review_column"]
    )

    headers = {
        match_type_column: "Match Type",
        partner_column: "Partner Rows",
        review_column: "Review",
    }

    for column, header in headers.items():
        cell = sheet.cell(
            row=HEADER_ROW,
            column=column,
        )

        cell.value = header
        cell.font = Font(bold=True)

    last_column = get_result_last_column(sheet)

    for row in range(
        START_ROW,
        sheet.max_row + 1,
    ):
        for column in range(
            1,
            last_column + 1,
        ):
            sheet.cell(
                row=row,
                column=column,
            ).fill = PatternFill()

        sheet.cell(
            row=row,
            column=match_type_column,
        ).value = None

        sheet.cell(
            row=row,
            column=partner_column,
        ).value = None

        sheet.cell(
            row=row,
            column=review_column,
        ).value = None


def write_records(sheet, records):
    """写入匹配关系并按规则着色。"""

    config = get_sheet_config(sheet.title)

    match_type_column = (
        config["match_type_column"]
    )

    partner_column = (
        config["partner_column"]
    )

    review_column = (
        config["review_column"]
    )

    for record in records:
        fill_row(
            sheet=sheet,
            row=record.row,
            fill=get_fill(record),
        )

        if not record.matched:
            continue

        sheet.cell(
            row=record.row,
            column=match_type_column,
        ).value = record.match_type

        sheet.cell(
            row=record.row,
            column=partner_column,
        ).value = ", ".join(
            str(row)
            for row in record.partners
        )

        sheet.cell(
            row=record.row,
            column=review_column,
        ).value = record.review_reason


def set_result_column_widths(sheet):
    """设置Key word和结果列宽。"""

    config = get_sheet_config(sheet.title)

    widths = {
        config["key_word_column"]: 24,
        config["match_type_column"]: 32,
        config["partner_column"]: 24,
        config["review_column"]: 28,
    }

    for column, width in widths.items():
        column_letter = (
            get_column_letter(column)
        )

        sheet.column_dimensions[
            column_letter
        ].width = width


def save_results(
    workbook,
    sheet1_records,
    sheet2_records,
    output_filename,
    difference_result=None,
):
    """把最终匹配结果写入工作簿。"""

    sheet1 = workbook["Sheet1"]
    sheet2 = workbook["Sheet2"]

    clear_old_results(sheet1)
    clear_old_results(sheet2)

    write_records(
        sheet=sheet1,
        records=sheet1_records,
    )

    write_records(
        sheet=sheet2,
        records=sheet2_records,
    )

    set_result_column_widths(sheet1)
    set_result_column_widths(sheet2)

    workbook.save(output_filename)